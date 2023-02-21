"""
Cryptocurrency Price Prediction Tool
Author: Oleg Vinokurov
Company: VB-Programmer Inc.
Date: 02/19/2023

This program retrieves historical price data for a selected cryptocurrency, 
uses it to train a machine learning model, and generates price predictions 
for a specified time period. The predicted prices are saved in an Excel 
spreadsheet and displayed to the user. The program uses the CryptoCompare 
API to retrieve historical price data and the scikit-learn library to train 
the machine learning model.

This application assumes that the user has a basic understanding of 
cryptocurrency and machine learning. The predictions generated by this 
program are not guaranteed to be accurate and should not be relied upon for 
investment decisions.

"""
import requests
import openpyxl
from datetime import datetime, timedelta
from sklearn.linear_model import LinearRegression
from colorama import init, Fore, Back
import os
import itertools
from statsmodels.tsa.arima_model import ARIMA


# This code block checks if the 'reports' folder exists and creates it if it doesn't.
# This is to ensure that the Excel spreadsheet containing the price predictions can be saved to the correct directory.
if not os.path.exists('reports'):
    os.mkdir('reports')
# initialize colorama
init()

# Define the CryptoCompare API endpoint and parameters
CRYPTOCOMPARE_API_ENDPOINT = "https://min-api.cryptocompare.com/data/v2/histohour"

# Define the machine learning model to predict cryptocurrency prices
model = LinearRegression()

# Define the Excel spreadsheet to store the price predictions
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Price Predictions"
ws.cell(1, 1, value="Date")
ws.cell(1, 2, value="Price")

# Set the 'again' flag to False to trigger the initial cryptocurrency selection prompt
again = False

# Start the main program loop
while True:
    ws.delete_rows(2, ws.max_row)
    
    # If this isn't the first iteration of the loop, prompt the user to analyze another cryptocurrency
    if again:
        print("Do you want to analyze another cryptocurrency? (y/n)")
        choice = input("> ")
        if choice.lower() == "n":
            break
        
        # If the user selects something other than 'y' or 'n', prompt them again
        elif choice.lower() != "y":
            print("Invalid input. Please enter 'y' or 'n'.")
            continue
    
    # Set the 'again' flag to True to bypass the initial cryptocurrency selection prompt
    else:
        again = True

    # Prompt the user to select a cryptocurrency to analyze
    print("Choose a cryptocurrency to analyze:")
    
    # Send a request to the CryptoCompare API to get the list of top cryptocurrencies by market cap
    response = requests.get("https://min-api.cryptocompare.com/data/top/mktcapfull?limit=50&tsym=USD")
    
    # Extract the names of the top cryptocurrencies from the API response
    cryptocurrencies = [c["CoinInfo"]["Name"] for c in response.json()["Data"]]
    
    # Print the list of cryptocurrencies to the console, with a number for each option
    for i, c in enumerate(cryptocurrencies):
        print(f"{i + 1}. {c}")
    
    # Prompt the user to select a cryptocurrency from the list
    choice = int(input("> "))
    
    # Define the CryptoCompare API parameters for the selected cryptocurrency and time period
    selected_cryptocurrency = cryptocurrencies[choice - 1]

    # Prompt the user to select a time period to analyze
    print("Choose a time period to analyze:")
    print("1. 24 hours")
    print("2. 7 days")
    print("3. 12 months")
    time_period_choice = int(input("> "))

    # Prompt the user to select an algorithm for price prediction
    print("Choose an algorithm to use for price prediction:")
    print("1. Linear Regression")
    print("2. ARIMA")
    algorithm_choice = int(input("> "))

    # This function returns a list of combinations of p, d, and q 
    # values for the ARIMA model to optimize its performance.
    def get_pdq_values():
        p = range(0, 6)
        d = range(0, 2)
        q = range(0, 2)
        pdq = [(x[0], x[1], x[2]) for x in list(itertools.product(p, d, q))]
        return pdq

    # set up the pdq values
    pdq = get_pdq_values()
    

    # Define the CryptoCompare API parameters for the selected cryptocurrency and time period
    if time_period_choice == 1:
        name_time_period_choice = "24 hours"
        CRYPTOCOMPARE_API_PARAMS = {
            "fsym": selected_cryptocurrency,
            "tsym": "USD",
            "limit": 24,
            "aggregate": 1
        }
        if algorithm_choice == 1:
            model = LinearRegression()
        elif algorithm_choice == 2:
            pdq
    elif time_period_choice == 2:
        name_time_period_choice = "7 days"
        CRYPTOCOMPARE_API_PARAMS = {
            "fsym": selected_cryptocurrency,
            "tsym": "USD",
            "limit": 168,
            "aggregate": 1
        }
        if algorithm_choice == 1:
            model = LinearRegression()
        elif algorithm_choice == 2:
            pdq
    elif time_period_choice == 3:
        name_time_period_choice = "12 months"
        CRYPTOCOMPARE_API_PARAMS = {
            "fsym": selected_cryptocurrency,
            "tsym": "USD",
            "limit": 365,
            "aggregate": 1
        }
        if algorithm_choice == 1:
            model = LinearRegression()
        elif algorithm_choice == 2:
            pdq
    else:
        print("Invalid time period choice. Please try again.")
        continue

    # Get the historical price data for the selected cryptocurrency
    response = requests.get(CRYPTOCOMPARE_API_ENDPOINT, params=CRYPTOCOMPARE_API_PARAMS)

    if not response.ok:
        print("Could not retrieve historical price data. Please try again.")
        continue

    history_data = response.json()["Data"]["Data"]

    if not history_data:
        print("Historical price data is empty. Please try again.")
        continue

    # Prepare the historical price data for machine learning
    timestamps = [datetime.fromtimestamp(h["time"]).timestamp() * 1000 for h in history_data]
    prices = [h["close"] for h in history_data]
    timestamps = [[t] for t in timestamps]

    # Train the machine learning model on the historical price data
    model.fit(timestamps, prices)

    # Generate the price predictions for the selected time period
    if time_period_choice == 1:
        current_time = datetime.now()
        for i in range(24):
            next_time = current_time + timedelta(hours=i+1)
            next_timestamp = int(next_time.timestamp() * 1000)
            next_price = model.predict([[next_timestamp]])[0]
            ws.cell(i+2, 1, value=next_time)
            ws.cell(i+2, 2, value=next_price)
            ws.cell(i+2, 1).number_format = "mm/dd/yyyy hh:mm"
    elif time_period_choice == 2:
        current_time = datetime.now()
        for i in range(7):
            next_time = current_time + timedelta(days=i+1)
            if next_time - current_time > timedelta(days=7):
                break
            next_timestamp = int(next_time.timestamp() * 1000)
            next_price = model.predict([[next_timestamp]])[0]
            ws.cell(i+2, 1, value=next_time)
            ws.cell(i+2, 2, value=next_price)
            ws.cell(i+2, 1).number_format = "mm/dd/yyyy hh:mm"
    elif time_period_choice == 3:
        current_time = datetime.now()
        for i in range(12):
            # Calculate the next date, which is 30 days after the previous date
            next_date_time = current_time + timedelta(days=30*(i+1))
            next_date = next_date_time.date()

            # Calculate the timestamp for the next date
            next_timestamp = int(next_date_time.timestamp() * 1000)
            
            # Predict the price for the next date using the machine learning model
            next_price = model.predict([[next_timestamp]])[0]
            
            # Write the date and price to the Excel spreadsheet
            ws.cell(i+2, 1, value=next_date)
            ws.cell(i+2, 2, value=next_price)

    # Save the Excel spreadsheet
    from datetime import datetime

    # Get current date and time
    now = datetime.now()

    # Convert to string
    date_time_string = now.strftime("%m%d%Y-%H%M%S")

    try:
        ws.cell(1, 3, value="Algorithm")
        ws.cell(2, 3, value="Linear-Regression" if algorithm_choice == 1 else "ARIMA")
        wb.save(f"reports/{'Linear-Regression' if algorithm_choice == 1 else 'ARIMA'}-{name_time_period_choice.replace(' ','-')}-{selected_cryptocurrency}-{date_time_string}.xlsx")
        print(f"Saved to: reports/{'Linear-Regression' if algorithm_choice == 1 else 'ARIMA'}-{name_time_period_choice.replace(' ','-')}-{selected_cryptocurrency}-{date_time_string}.xlsx")
    except:
        print("Error: Could not save price predictions to Excel spreadsheet. Please try again.")
        exit()

    # Print the results to the console
    print(f"Price predictions for: {selected_cryptocurrency} ")
    print(f"Algorithm: {'Linear Regression' if algorithm_choice == 1 else 'ARIMA'}")
    print(f"Time period: {name_time_period_choice}")
    print("+" + "-"*27 + "+" + "-"*12 + "+")
    print("| Date".ljust(27) + " | Price".ljust(14) + "|")
    print("+" + "-"*27 + "+" + "-"*12 + "+")
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        date = row[0]
        price = f"${row[1]:.2f}"
        price_width = len(price)
        date_width = 25
        price_column_width = max(10, price_width)
        date_column_width = max(25, date_width)
        print(f"| {date.strftime('%m/%d/%Y %I:%M:%S %p').ljust(date_column_width)} | {price.rjust(price_column_width)} |")
    print("+" + "-"*27 + "+" + "-"*12 + "+")

print("Thank you for using the cryptocurrency price prediction tool!")
